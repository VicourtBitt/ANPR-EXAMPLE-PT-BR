import pandas as pd
from simpledt import ExcelDataTable
import flet as flt
import time 
import datetime
import paho.mqtt.publish as publish
from openpyxl import load_workbook

def main(page= flt.Page):
    page.title = "Sistema de Reconhecimento de Placas de Veiculares"
    page.window_width = 580
    page.window_max_width = 580
    page.window_height = 600
    page.window_max_height = 600
    page.scroll = "auto"
    page.horizontal_alignment = "Center"
    page.padding = 5

    MQTT_TOPIC = "placasdecar1"
    MQTT_BROKER = "test.mosquitto.org"
    MQTT_PORT = 1883

    excel_path = r"C:\\Users\\Victor\\Documents\\NewSheet.xlsx"
    platepos = 0
    sheet_name1 = "Planilha1"

    def route_change(route):
        
        # ———————————————————————————————————————————————————————————————————————————————————————————————————————
        #                                          Declaração Insana das Respostas
        # ———————————————————————————————————————————————————————————————————————————————————————————————————————

        #——————————————————————————————————————————————————————————

        def EspCicle():
            LowEsp()
            HighEsp()
            time.sleep(40.0)
            LowEsp()
            HighEsp()
            LowEsp()
        
        def HighEsp():
            publish.single(MQTT_TOPIC, payload="1", qos=0, retain=False, 
                                hostname=MQTT_BROKER, port=MQTT_PORT)
            time.sleep(0.5)

        def LowEsp():
            publish.single(MQTT_TOPIC, payload="0", qos=0, retain=False, 
                                hostname=MQTT_BROKER, port=MQTT_PORT)
            time.sleep(0.5)

        #——————————————————————————————————————————————————————————
        
        def openingCicle(e):
            Z1Button.data += 1
            responseZ1Button.value = f'Ciclo Acionado {Z1Button.data} vez(es).'
            page.update()
            EspCicle()
            page.update()

        def pulseCicle(e):
            Z2Button.data += 1
            responseZ2Button.value = f'Pulso Acionado {Z2Button.data} vez(es).'
            page.update()
            LowEsp()
            HighEsp()
            LowEsp()
        
        #——————————————————————————————————————————————————————————

        def actionRegister(e):
            global platepos
            platepos = 0
            global dfPandas
            dfPandas = pd.read_excel(excel_path)
            plateR = dfPandas.iloc[platepos, 0]
            wb = load_workbook(excel_path)
            ws = wb[sheet_name1]

            Z3Button.value = f'A placa que foi recém cadastrada é: {Z1TextField.value}'
            while True:
                dfPandas = pd.read_excel(excel_path)
                lenghtDF = (len(dfPandas["Dauri"].index))

                if (len(Z1TextField.value) < 7):
                    Z3Button.value = f"TENTATIVA NEGADA PLACA COM MENOS DE 7 DÍGITOS"
                    break

                if (platepos == lenghtDF) and (Z1TextField.value not in dfPandas):
                    wb = load_workbook(excel_path)
                    ws = wb[sheet_name1]
                    plate_format = f'{Z1TextField.value:.7}'
                    new_row = [plate_format]
                    ws.append(new_row)
                    wb.save(excel_path)
                    break

                plateR = dfPandas.iloc[platepos, 0]
                if platepos < lenghtDF:
                    if Z1TextField.value != plateR:
                        platepos += 1
                    else:
                        Z3Button.value = f" A PLACA {Z1TextField.value} JÁ EXISTE NO BANCO DE DADOS."
                        break
            
            responseZ3Button.value = Z3Button.value
            page.update()

        # ———————————————————————————————————————————————————————————————————————————————————————————————————————
        #                                          Declaração Insana da Tela Inicial
        # ———————————————————————————————————————————————————————————————————————————————————————————————————————

        responseZ1Button = flt.Text(size=22)
        separator = flt.Text(" | ", size=23)
        responseZ2Button = flt.Text(size=22)

        global responseZ3Button
        responseZ3Button = flt.Text(size=22)

        Z1column = flt.Column()
        Z0row = flt.Row()
        Z1row = flt.Row()
        Z1d5row = flt.Row()
        Z2row = flt.Row()
        Z2d5row = flt.Row()

        Z0CopyWrite = flt.Text(
            "Aplicação Digital feita por Victor Bittencourt em vigência do TCC Projeto Integrador SENAI/FIERGS.",
            size=12,
            italic=True,
            color= flt.colors.GREY_100,
            )

        plate_container = flt.Container(
            content= flt.Image(
                                    src= r'Python\\ProjetoSENAI\\CursoPython\\Markus\\Imagem\\plate_0.jpg',
                                    width= 300,
                                    height = 97,
                                    scale=1.75,
                                    repeat= flt.ImageRepeat.REPEAT,
                                    fit= flt.ImageFit.FIT_WIDTH,
                                    gapless_playback= False,
                                ),
                                alignment= flt.alignment.top_center,
                                padding= 40
                            )  
        
        #——————————————————————————————————————————————————————————

        Z1Button = flt.ElevatedButton(
                        "Ciclo de Abertura do Portão",
                        data=0,
                        on_click= openingCicle,
                        width=270,
                        height=35,
                        style= flt.ButtonStyle(
                                            shape={
        flt.MaterialState.DEFAULT: flt.RoundedRectangleBorder(radius=3)
                                                }
                                            ),
                                         )
        
        Z1ButtonContainer = flt.Container(
        content = Z1Button,
        padding= 0
                                    )
        
        #——————————————————————————————————————————————————————————

        Z2Button = flt.ElevatedButton(
                        "Pulso Manual de Abertura",
                        data= 0,
                        on_click= pulseCicle,
                        width=270,
                        height=35,
                        style= flt.ButtonStyle(
                                            shape={
        flt.MaterialState.DEFAULT: flt.RoundedRectangleBorder(radius=3)
                                                }
                                            ),
                                        )

        Z2ButtonContainer = flt.Container(
        content= Z2Button,
        padding= 0
                                    )
        
        #——————————————————————————————————————————————————————————

        Z1TextField = content= flt.TextField(
        width= 350,
        height= 80,
        label= "Cadastre ou Verifique a Placa no sistema",
        hint_text="Escreva somente 7 digitos, ignore hífens.",
        max_length= 7
                            )

        Z1TextFieldContainer = flt.Container(
        content= Z1TextField,
        padding= 0
                            )                   
        
        #——————————————————————————————————————————————————————————

        Z3Button = flt.ElevatedButton(
                        on_click=actionRegister,
                        text="Cadastre/Verifique",
                        width=190,
                        height=80,
                        style= flt.ButtonStyle(
                                            shape={
        flt.MaterialState.DEFAULT: flt.ContinuousRectangleBorder(radius=0)
                                                }
                                            )
                                        )

        Z3ButtonContainer = flt.Container(
        content = Z3Button,
        padding=0
                                        )

        #——————————————————————————————————————————————————————————

        Z1column.controls.append(plate_container)

        Z0row.controls.append(Z0CopyWrite)

        Z1row.controls.append(Z1ButtonContainer)
        Z1row.controls.append(Z2ButtonContainer)

        Z1d5row.controls.append(responseZ1Button)
        Z1d5row.controls.append(separator)
        Z1d5row.controls.append(responseZ2Button)

        Z2row.controls.append(Z1TextFieldContainer)
        Z2row.controls.append(Z3ButtonContainer)

        Z2d5row.controls.append(responseZ3Button)

        Z1column.controls.append(Z0row)
        Z1column.controls.append(Z1row)
        Z1column.controls.append(Z1d5row)
        Z1column.controls.append(Z2row)
        Z1column.controls.append(Z2d5row)

        # ———————————————————————————————————————————————————————————————————————————————————————————————————————
        #                                             TELA INICIAL
        # ———————————————————————————————————————————————————————————————————————————————————————————————————————

        page.views.clear()
        page.views.append(
            flt.View(
                '/',
                [
                    flt.AppBar(
                        title=flt.Text("TELA INICIAL"),
                        center_title=True,
                        toolbar_height= 50, 
                        bgcolor=flt.colors.BLUE_900,
                        actions= 
                                [
                                flt.IconButton(flt.icons.HISTORY, on_click=lambda _: page.go('/page2'), icon_size=30),
                                flt.IconButton(flt.icons.LIST_ALT, on_click=lambda _: page.go('/page1'), icon_size=30)
                                ]
                            ),
                    Z1column
                ], page.update()
            )
        )
        page.update()
        # ———————————————————————————————————————————————————————————————————————————————————————————————————————
        #                                   Declaração Insada das Cadastradas
        # ———————————————————————————————————————————————————————————————————————————————————————————————————————

        H0Column = flt.Column()
        H0Row = flt.Row()
        H1Row = flt.Row()

        excel_path2 = r"C:\\Users\\Victor\\Documents\\teste.xlsx"
        dataFrame2 = ExcelDataTable(excel_path2)
        dataTable2 = dataFrame2.datatable

        ContainerDT2 = flt.Container(
            content= dataTable2,
            width=560
                                )
        
        H0CopyWrite = flt.Text(
            "Aplicação Digital feita por Victor Bittencourt em vigência do TCC Projeto Integrador SENAI/FIERGS.",
            size=12,
            italic=True,
            color= flt.colors.GREY_100,
            )
        
        H0Row.controls.append(H0CopyWrite)
        H1Row.controls.append(ContainerDT2)

        H0Column.controls.append(H0Row)
        H0Column.controls.append(H1Row)

        # ———————————————————————————————————————————————————————————————————————————————————————————————————————
        #                                   Declaração Insada das Cadastradas
        # ———————————————————————————————————————————————————————————————————————————————————————————————————————

        R0column = flt.Column()
        R0row = flt.Row()
        R1row = flt.Row()

        excel_path = r"C:\\Users\\Victor\\Documents\\dauri.xlsx"
        dataFrame = ExcelDataTable(excel_path)
        dataTable = dataFrame.datatable

        ContainerDT = flt.Container(
            content= dataTable,
            width=560,
                                )

        R0CopyWrite = flt.Text(
            "Aplicação Digital feita por Victor Bittencourt em vigência do TCC Projeto Integrador SENAI/FIERGS.",
            size=12,
            italic=True,
            color= flt.colors.GREY_100,
            )

        R0row.controls.append(R0CopyWrite)

        R1row.controls.append(ContainerDT)

        R0column.controls.append(R0row)
        R0column.controls.append(R1row)

        # ———————————————————————————————————————————————————————————————————————————————————————————————————————
        #                                           TELA DAS CADASTRADAS
        # ———————————————————————————————————————————————————————————————————————————————————————————————————————

        if page.route == "/page1":
            page.views.append(
                flt.View(
                '/page1',
                [
                    flt.AppBar(
                        title=flt.Text("Placas Cadastradas"),
                        center_title=True,
                        toolbar_height= 50, 
                        bgcolor=flt.colors.BLUE_GREY_800,
                        actions= 
                                [
                                flt.IconButton(flt.icons.HISTORY, on_click=lambda _: page.go('/page2'), icon_size=30),
                                flt.IconButton(flt.icons.HOME, on_click=lambda _: page.go('/'), icon_size=30)
                                ]
                            ),
                   R0column, 
                ], page.update(),            
                scroll= "auto",
            )
        )
            
            
        # ———————————————————————————————————————————————————————————————————————————————————————————————————————
        #                                          TELA DO HISTÓRICO
        # ———————————————————————————————————————————————————————————————————————————————————————————————————————    

        elif page.route == "/page2":
            page.views.append(
                flt.View(
                '/page2',
                [
                    flt.AppBar(
                        title=flt.Text("Histórico"),
                        center_title=True,
                        toolbar_height= 50, 
                        bgcolor=flt.colors.GREY_800,
                        actions=
                                [
                                flt.IconButton(flt.icons.HOUSE, on_click=lambda _: page.go('/'), icon_size=30),
                                flt.IconButton(flt.icons.LIST_ALT, on_click=lambda _: page.go('/page1'), icon_size=30)
                                ]
                            ),
                   H0Column
                ],
            )
        )
        page.update()

    def view_pop(view):
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    page.on_route_change = route_change
    page.on_view_pop = view_pop
    page.go(page.route)

    page.update()

flt.app(target=main)
